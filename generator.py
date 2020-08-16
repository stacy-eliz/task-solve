import classes
from parser2 import application2_audit, application2_programm
from parser1 import application1
from reading_teachers import Full_teacher
from random import shuffle
from openpyxl import load_workbook



wb = load_workbook("output.xlsx")
ws1 = wb["ДПО"]

########Запись ячейки в файл ########
def room_xls(l, s, st):
    if s != 0:  
        x = st.find("ауд.")
        audit = st[:x]+ " ауд. " + str(s.name)+st[x+4:]
        ws1.cell(row=l[1] + 1, column=l[0] + 7, value=audit)
####################


###Сохранение измененного файла#######
def save():
    wb.save("output.xlsx")
#######################3

Shedule1 = classes.Shedule()
path_app1 = "app1.xlsx"
##path_app1 = "output.xlsx"
a = application1(path_app1)
path = "app2.xlsx"

b = application2_programm(path)  # массив объектов программ
c = application2_audit(path)  # массив объектов аудиторий
s = 0

counter_weeks = 0
count_free_room_weeks = 0

count_free_room_week = 0

y1 = []
y2 = []

y3 = []
y4 = []


##########Функция обрезания строки#############
def shorter(st, f):
    if f:
        for i in range(len(st)):
            if st[i].isnumeric():
                return st[:i]
    else:
        for i in range(len(st)):
            if st[i].isnumeric():
                return st[i:][:11]
    return st
###############################################

teachers = Full_teacher('app2.xlsx') ##чтение фала "Приложение№2" и заполнение массива учителей

######################Составление расписания учителей#########################
flag = False
flag2 = False
count = 0
unic = []
for i in range(1, len(a)):
    flag1 = True
    for j in range(len(a[i])):
        if flag1 and a[i][j] != None and 'Неделя' in a[i][j]:
            count += 1
            flag1 = False
        elif a[i][j] != None and not 'Неделя' in a[i][j]:
            if a[i][j] not in unic:
                unic.append(a[i][j])
                if isinstance(a[i][j], list):
                    # print(a[i][j], count)
                    for k in a[i][j]:
                        shuffle(teachers)
                        for m in teachers:
                            x = m.disciplin.replace('Управление безопасностью полетов', 'Безопасность полетов').replace(
                                'Организация пассажирских перевозок', 'Пассажирские перевозки')
                            if x in a[0][j] or a[0][j] in x:
                                if not m.is_busy:
                                    Shedule1.Add_Teacher(m, shorter(k, 0), x, k, count)
                                    m.is_busy = 1
                                    flag2 = True
                                    break
                        if flag2:
                            flag2 = False
                            break
                else:
                    shuffle(teachers)
                    for m in teachers:
                        x = m.disciplin.replace('Управление безопасностью полетов', 'Авиационная безопасность').replace(
                            'Организация пассажирских перевозок', 'Пассажирские перевозки')
                        if x in a[0][j] or a[0][j] in x:
                            if not m.is_busy:
                                Shedule1.Add_Teacher(m, shorter(a[i][j], 0), x, a[i][j], count)
                                m.is_busy = 1
                                flag = True
                                break
                    if flag:
                        flag = False
                        break
    Shedule1.leave_teachers()

del unic
##########################################################################


###################Составление расписания, анализ данных о парах в месяц/неделю###################################
for i in range(1, len(a)):
    counter_weeks += 1
    for j in range(len(a[i])):
        if not (type(a[i][j]) is list) and a[i][j] != None and a[0][j] != None:
            for k in range(len(b)):
                if "Опасные грузы" in a[0][j]:
                    cate = a[0][j].replace("Опасные грузы.", "").replace(" категория", "")
                    if "опасные грузы" in b[k].name.lower():
                        cate += " категория ИКАО/ИАТА"
                        if "Базовый" in a[i][j]:
                            cate += ".базовый курс"
                        if cate.lower() in b[k].programme.lower() or b[k].programme.lower() in cate.lower():
                            f = Shedule1.get_free_rooms()
                            if len(f) != 0:
                                for s in f:
                                    s.is_busy = 1
                                    Shedule1.Add_Room(s, a[i][2], None, b[k].programme, [i, j])
                                    room_xls([i, j], s, a[i][j])
                                    break


                else:
                    d = shorter(a[i][j], True).replace("Повышение", "повышения")
                    d = d.lower()
                    if a[0][j].lower() in b[k].name.lower() or b[k].name.lower() in a[0][j].lower():
                        if d in b[k].programme.lower() or b[k].programme.lower() in d:
                            for l in range(len(c)):
                                h = b[k].features.lower().replace("требуется ", "").replace("нет", "")
                                if str(c[l].name) in h:
                                    if c[l].is_busy:
                                        f = Shedule1.get_free_rooms()
                                        if len(f) != 0:
                                            for s in f:
                                                s.is_busy = 1
                                                Shedule1.Add_Room(s, a[i][2], None, b[k].programme, [i, j])
                                                Shedule1.swap_rooms(c[l], s, a[i][2])
                                                room_xls([i, j], s, a[i][j])
                                                break
                                    else:
                                        Shedule1.Add_Room(c[l], a[i][2], None, b[k].programme, [i, j])
                                        room_xls([i, j], s, a[i][j])
                                elif h in c[l].differences.lower() and not (c[l].is_busy):

                                    Shedule1.Add_Room(c[l], a[i][2], None, b[k].programme, [i, j])
                                    room_xls([i, j], s, a[i][j])
                                    c[l].is_busy = 1
                                    break
        elif type(a[i][j]) is list:
            for g in range(len(a[i][j])):
                for k in range(len(b)):
                    d = shorter(a[i][j][g], True).replace("Повышение", "повышения")
                    d = d.lower()
                    if a[0][j][g].lower() in b[k].name.lower() or b[k].name.lower() in a[0][j][g].lower():
                        if d in b[k].programme.lower() or b[k].programme.lower() in d:
                            for l in range(len(c)):
                                h = b[k].features.lower().replace("требуется ", "").replace("нет", "")
                                if str(c[l].name) in h:
                                    if c[l].is_busy:
                                        f = Shedule1.get_free_rooms()
                                        if len(f) != 0:
                                            for s in f:
                                                s.is_busy = 1
                                                Shedule1.Add_Room(s, a[i][2], None, b[k].programme, [i, j])
                                                Shedule1.swap_rooms(c[l], s, a[i][2])
                                                room_xls([i, j], s, a[i][j][g])
                                                break
                                    else:
                                        Shedule1.Add_Room(c[l], a[i][2], None, b[k].programme, [i, j])
                                        room_xls([i, j], s, a[i][j][g])
                                elif h in c[l].differences.lower() and not (c[l].is_busy):
                                    Shedule1.Add_Room(c[l], a[i][2], None, b[k].programme, [i, j])
                                    room_xls([i, j], s, a[i][j][g])
                                    c[l].is_busy = 1
                                    break
    ######Анализ данных для графиков##########
    if i % 2 == 0:  
        count_free_room_week = len(Shedule1.get_free_rooms())
        y2.append(count_free_room_week / 11 * 100)

    count_free_room_weeks += len(Shedule1.get_free_rooms())
    if counter_weeks == 4:
        y1.append(count_free_room_weeks / 44 * 100)
        counter_weeks = 0
        count_free_room_weeks = 0
    #################

    Shedule1.leave_rooms()
###################################################


#############Анализ данных для графиков##################
teacher_time = {}
cons = 1
for j in Shedule1.busy_teachers:
    n = j[0]  # имя препода
    w = Shedule1.busy_teachers[j][-1]  # номер недели
    a = Shedule1.busy_teachers[j][-2]
    a = a.split()
    if a[2][3:5] != a[2][-2:]:
        cons += int(a[2][6:8])
        cons += 30 - int(a[2][:2])
        cons += 1
    else:
        cons = int(a[2][6:8]) - int(a[2][:2]) + 1
    if n in teacher_time:
        if w in teacher_time[n]:
            teacher_time[n][w] += cons
        else:
            teacher_time[n][w] = cons
    else:
        teacher_time[n] = {w: cons}
#############################
save()
