from parser1 import application1
from generator import Shedule1
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("output.xlsx")

path_app1 = "app1.xlsx"
a = application1(path_app1)
s = ""
ws1 = wb.get_sheet_by_name("ДПО")
for i in range(1, len(a)):
    for j in range(len(a[i])):
        
       
        if a[i][j] != None and not("Неделя" in a[i][j]):
            if isinstance(a[i][j], list):
##                    print("&&&", a[i][j])
                    for k in a[i][j]:
                        s+= " "
##                        print("&&&&",k)
                        s += k
                        
                        for m in Shedule1.busy_teachers:
                            if Shedule1.busy_teachers[m][1] in k and Shedule1.busy_teachers[m][0] in a[0][j]:
##                                print(type(m[0]))
                                print("find")
                                s+=" "
                                s+=m[0]
##                                print("*", Shedule1.busy_teachers[m][1])
##                                s=s+" "+str(Shedule1.busy_teachers[m][1] )
                                
                                break
                    print(s)
                                
                    ws1.cell(row=j+1, column=i+7,value=s)
                    s = ""
                        
            else:
                s+=""
                s += a[i][j]
##                print(a[i][j], a[0][j])
                for m in Shedule1.busy_teachers:
                    if Shedule1.busy_teachers[m][1] in a[i][j] and Shedule1.busy_teachers[m][0] in a[0][j]:
                        s+=str(m[0])
##                        s=s+" "+Shedule1.busy_teachers[m][1]
                        print("*", s)
                        break
                        
                ws1.cell(row=j+1, column=i+7,value=s)
                s = ''
                        
wb.save("output.xlsx")
                        
            
            
