import openpyxl


def application1(path):
    wb = openpyxl.load_workbook(path)  # 'Приложение №1.xlsx'

    direction = []
    weeks = []
    sheet = wb['ДПО']
    for column in sheet.iter_cols(6, 7):
        for j in column:
            x = j.value
            if x != None:
                direction.append(str(x).replace('\n', ' ').strip())
            else:
                direction.append(None)

    for column in sheet.iter_cols(8, sheet.max_column):
        weeks.append([])
        for j in column:
            if j.value != None:
                x = str(j.value).replace('\n', ' ').strip()
                if not x.isupper() and x[0].isalpha():
                    if x.count('ауд.') == 1:
                        y = x.find('ауд.')
                        x = x[:y + 4]
                        weeks[-1].append(x)
                    elif x.count('ауд.') > 1:
                        k = x.split()
                        x = ['']
                        flag = False
                        for m in k:
                            if m == 'ауд.':
                                flag = True
                                x[-1] += m + ' '
                                x.append('')
                            elif flag:
                                flag = False
                            else:
                                x[-1] += m + ' '
                        for m in range(len(x) - 1):
                            x[m] = x[m].strip()
                        weeks[-1].append(x[:-1])
                else:
                    weeks[-1].append(None)
            else:
                weeks[-1].append(None)
        if weeks[-1] == [] or weeks[-1] == ['Неделя 52 23.12-29.12'] or weeks[-1] == ['Неделя 01 30.12-05.01']:
            weeks.pop()
    weeks.insert(0, direction)
    weeks[0] = weeks[0][42:]
    return weeks
