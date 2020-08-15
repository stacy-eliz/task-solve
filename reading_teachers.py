import openpyxl
from pathlib import Path
from classes import Teacher
def Full_teacher(path) -> [Teacher]:
  xlsx_file = Path(path)
  wb_obj = openpyxl.load_workbook(xlsx_file)
  sheet = wb_obj['параметры преподавателей']
  pr = []
  NUMBERS_OF_PROGRAMS = []
  for row in sheet.iter_rows(min_row=2):
    pr.append(str(row[3].value).split(';'))
  flatten = lambda l: [item for sublist in l for item in sublist]
  for i in list(set(flatten(pr))):
    NUMBERS_OF_PROGRAMS.append(i)
  del pr
  teachers = []
  for _, row in enumerate(sheet.iter_rows(min_row=2)):
    azaza = []
    [azaza.append(j) for j in NUMBERS_OF_PROGRAMS if j not in str(row[3].value).split(';')]
    tch = Teacher(row[0].value, row[1].value, str(row[3].value).split(';'), azaza, row[4].value, row[7].value, row[2].value)
    teachers.append(tch)
  return teachers